import openpyxl


excel_file = openpyxl.load_workbook('./avtovoz_cars.xlsx')
sheet = excel_file.active


excel_file = openpyxl.load_workbook('./avtovoz_cars.xlsx')
marka_from_file = input("marka:")

sheet = excel_file.active


for row in range(1,sheet.max_row+1):
    marka = sheet[row][0].value
    model = sheet[row][1].value
    type = sheet[row][2].value
    system = sheet[row][3].value
    price = sheet[row][4].value
    transporter = sheet[row][5].value
    dimensions = sheet[row][6].value

    if marka == marka_from_file:
        print(model, type, system, price, transporter, dimensions)
